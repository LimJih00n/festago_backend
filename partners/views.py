from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Q, Sum, Avg
from django.http import HttpResponse
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .models import Partner, Application, Message, AnalyticsData, ImageUpload, Notification, ApplicationDraft, FestivalBookmark
from events.models import Event
from .serializers import (
    PartnerSignupSerializer,
    PartnerSerializer,
    PartnerPublicSerializer,
    ApplicationSerializer,
    ApplicationCreateSerializer,
    MessageSerializer,
    AnalyticsDataSerializer,
    ImageUploadSerializer,
    NotificationSerializer,
    ApplicationDraftSerializer,
    FestivalBookmarkSerializer
)
from .utils import resize_image
from .pdf_generator import generate_analytics_pdf


class IsPartner(permissions.BasePermission):
    """사업자 권한 체크"""
    def has_permission(self, request, view):
        return request.user.is_authenticated and request.user.user_type == 'partner'


class PartnerSignupView(APIView):
    """파트너 회원가입 View"""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = PartnerSignupSerializer(data=request.data)
        if serializer.is_valid():
            result = serializer.save()
            return Response({
                'message': '파트너 회원가입이 완료되었습니다.',
                'user_id': result['user'].id,
                'partner_id': result['partner'].id,
                'username': result['user'].username,
                'email': result['user'].email,
                'brand_name': result['partner'].brand_name
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class PartnerViewSet(viewsets.ModelViewSet):
    """사업자 프로필 ViewSet"""
    queryset = Partner.objects.all()
    serializer_class = PartnerSerializer

    def get_permissions(self):
        if self.action in ['list', 'retrieve', 'public']:
            return [permissions.AllowAny()]
        return [IsPartner()]

    @action(detail=False, methods=['get'])
    def me(self, request):
        """내 프로필 조회"""
        try:
            partner = request.user.partner_profile
            serializer = self.get_serializer(partner)
            return Response(serializer.data)
        except Partner.DoesNotExist:
            return Response(
                {'detail': '사업자 프로필이 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['get'], serializer_class=PartnerPublicSerializer)
    def public(self, request, pk=None):
        """공개 프로필 조회 (브랜드 페이지)"""
        partner = self.get_object()
        serializer = PartnerPublicSerializer(partner)
        return Response(serializer.data)


class ApplicationViewSet(viewsets.ModelViewSet):
    """축제 지원서 ViewSet"""
    queryset = Application.objects.all()
    permission_classes = [IsPartner]

    def get_serializer_class(self):
        if self.action == 'create':
            return ApplicationCreateSerializer
        return ApplicationSerializer

    def get_queryset(self):
        # 자기 자신의 지원서만 조회
        return Application.objects.filter(
            partner__user=self.request.user
        ).select_related('partner', 'event')

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """지원서 승인"""
        application = self.get_object()
        organizer_message = request.data.get('organizer_message', '')
        application.approve(organizer_message)
        serializer = self.get_serializer(application)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """지원서 거절"""
        application = self.get_object()
        rejection_reason = request.data.get('rejection_reason', '')
        application.reject(rejection_reason)
        serializer = self.get_serializer(application)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """지원 내역 통계"""
        partner = request.user.partner_profile
        applications = partner.applications.all()

        stats = {
            'total': applications.count(),
            'pending': applications.filter(status='pending').count(),
            'approved': applications.filter(status='approved').count(),
            'rejected': applications.filter(status='rejected').count(),
            'completed': applications.filter(status='completed').count(),
        }
        return Response(stats)

    @action(detail=False, methods=['get'], url_path='export-excel')
    def export_excel(self, request):
        """지원서 엑셀 다운로드"""
        partner = request.user.partner_profile
        applications = self.get_queryset().select_related('event')

        # 워크북 생성
        wb = Workbook()
        ws = wb.active
        ws.title = "Applications"

        # 헤더 스타일
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # 헤더 작성
        headers = [
            'ID', 'Event Name', 'Status', 'Booth Type', 'Booth Size',
            'Products', 'Price Range', 'Applied Date', 'Reviewed Date',
            'Participation Fee', 'Payment Status', 'Booth Location'
        ]

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # 데이터 작성
        for row_num, app in enumerate(applications, 2):
            ws.cell(row=row_num, column=1, value=app.id)
            ws.cell(row=row_num, column=2, value=app.event.name)
            ws.cell(row=row_num, column=3, value=app.get_status_display())
            ws.cell(row=row_num, column=4, value=app.get_booth_type_display())
            ws.cell(row=row_num, column=5, value=app.get_booth_size_display())
            ws.cell(row=row_num, column=6, value=app.products)
            ws.cell(row=row_num, column=7, value=app.price_range)
            ws.cell(row=row_num, column=8, value=app.applied_at.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=9, value=app.reviewed_at.strftime('%Y-%m-%d') if app.reviewed_at else '')
            ws.cell(row=row_num, column=10, value=float(app.participation_fee))
            ws.cell(row=row_num, column=11, value=app.get_payment_status_display())
            ws.cell(row=row_num, column=12, value=app.booth_location)

        # 열 너비 자동 조정
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15

        # 파일 생성
        from io import BytesIO
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # 파일명 생성
        filename = f"applications_{partner.id}_{datetime.now().strftime('%Y%m%d')}.xlsx"

        # HTTP Response
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response


class MessageViewSet(viewsets.ModelViewSet):
    """메시지 ViewSet"""
    queryset = Message.objects.all()
    serializer_class = MessageSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        return Message.objects.filter(
            Q(sender=user) | Q(receiver=user)
        ).select_related('sender', 'receiver', 'application')

    def perform_create(self, serializer):
        serializer.save(sender=self.request.user)

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        """메시지 읽음 처리"""
        message = self.get_object()
        message.mark_as_read()
        serializer = self.get_serializer(message)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def inbox(self, request):
        """받은편지함"""
        messages = self.get_queryset().filter(receiver=request.user)
        serializer = self.get_serializer(messages, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def sent(self, request):
        """보낸편지함"""
        messages = self.get_queryset().filter(sender=request.user)
        serializer = self.get_serializer(messages, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        """읽지 않은 메시지 개수"""
        count = self.get_queryset().filter(
            receiver=request.user,
            read=False
        ).count()
        return Response({'unread_count': count})


class AnalyticsViewSet(viewsets.ReadOnlyModelViewSet):
    """성과 데이터 ViewSet (읽기 전용)"""
    queryset = AnalyticsData.objects.all()
    serializer_class = AnalyticsDataSerializer
    permission_classes = [IsPartner]

    def get_queryset(self):
        """
        현재 파트너의 성과 데이터를 반환
        데이터가 없으면 다른 파트너의 데이터를 랜덤하게 반환 (mock data)
        """
        try:
            partner = self.request.user.partner_profile
        except Partner.DoesNotExist:
            return AnalyticsData.objects.none()

        # 현재 파트너의 실제 데이터
        own_data = AnalyticsData.objects.filter(
            partner=partner
        ).select_related('partner', 'event', 'application')

        # 실제 데이터가 있으면 그대로 반환
        if own_data.exists():
            return own_data

        # 데이터가 없으면 다른 파트너의 데이터를 랜덤하게 가져옴 (mock data)
        # 최대 5개
        mock_data = AnalyticsData.objects.exclude(
            partner=partner
        ).select_related('partner', 'event', 'application').order_by('?')[:5]

        return mock_data

    @action(detail=True, methods=['get'], url_path='export-pdf')
    def export_pdf(self, request, pk=None):
        """PDF 리포트 다운로드"""
        analytics = self.get_object()
        partner_name = analytics.partner.brand_name
        event_name = analytics.event.name

        # PDF 생성
        pdf_buffer = generate_analytics_pdf(analytics, partner_name, event_name)

        # 파일명 생성
        filename = f"analytics_{analytics.partner.id}_{analytics.event.id}_{analytics.generated_at.strftime('%Y%m%d')}.pdf"

        # HTTP Response
        response = HttpResponse(pdf_buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """전체 성과 데이터 요약"""
        try:
            partner = request.user.partner_profile
        except Partner.DoesNotExist:
            return Response({'error': '파트너 프로필이 없습니다.'}, status=status.HTTP_404_NOT_FOUND)

        # 현재 파트너의 실제 데이터 확인
        own_data = AnalyticsData.objects.filter(partner=partner)
        is_mock_data = not own_data.exists()

        # get_queryset()이 mock data를 반환할 수 있음
        analytics_qs = self.get_queryset()

        # 집계 데이터
        summary_data = analytics_qs.aggregate(
            total_visitors=Sum('visitor_count'),
            total_sales=Sum('estimated_sales'),
            avg_rating=Avg('average_rating'),
            total_reviews=Sum('review_count'),
            avg_sentiment=Avg('sentiment_score'),
        )

        # 이벤트 수
        summary_data['total_events'] = analytics_qs.count()

        # mock data 여부 표시
        summary_data['is_sample_data'] = is_mock_data

        # 최고 성과 이벤트
        best_event = analytics_qs.order_by('-visitor_count').first()
        if best_event:
            summary_data['best_event'] = {
                'id': best_event.event.id,
                'name': best_event.event.name,
                'visitor_count': best_event.visitor_count,
                'sales': float(best_event.estimated_sales),
            }

        # 최근 3개 이벤트
        recent_analytics = analytics_qs.order_by('-generated_at')[:3]
        summary_data['recent_events'] = AnalyticsDataSerializer(recent_analytics, many=True).data

        return Response(summary_data)


class DashboardView(APIView):
    """대시보드 요약 데이터"""
    permission_classes = [IsPartner]

    def get(self, request):
        try:
            partner = request.user.partner_profile
        except Partner.DoesNotExist:
            return Response(
                {'detail': '사업자 프로필이 없습니다.'},
                status=status.HTTP_404_NOT_FOUND
            )

        # 통계
        applications = partner.applications.all()
        stats = {
            'pending': applications.filter(status='pending').count(),
            'approved': applications.filter(status='approved').count(),
            'completed': applications.filter(status='completed').count(),
            'total': applications.count(),
        }

        # 최근 알림
        notifications = []
        recent_messages = Message.objects.filter(
            receiver=request.user,
            read=False
        ).order_by('-created_at')[:5]

        for msg in recent_messages:
            notifications.append({
                'id': msg.id,
                'type': 'message',
                'message': msg.subject,
                'created_at': msg.created_at,
            })

        # 최근 승인/거절된 지원서
        recent_applications = applications.filter(
            status__in=['approved', 'rejected']
        ).order_by('-reviewed_at')[:3]

        for app in recent_applications:
            notifications.append({
                'id': app.id,
                'type': 'approval' if app.status == 'approved' else 'rejection',
                'message': f"{app.event.name} 지원이 {'승인' if app.status == 'approved' else '거절'}되었습니다.",
                'created_at': app.reviewed_at,
            })

        # 다가오는 일정
        upcoming_events = []
        approved_apps = applications.filter(
            status='approved',
            event__start_date__gte=date.today()
        ).select_related('event').order_by('event__start_date')[:5]

        for app in approved_apps:
            d_day = (app.event.start_date - date.today()).days
            upcoming_events.append({
                'id': app.event.id,
                'name': app.event.name,
                'date': app.event.start_date,
                'd_day': d_day,
                'location': app.event.location,
                'status': app.status,
            })

        data = {
            'stats': stats,
            'notifications': sorted(notifications, key=lambda x: x['created_at'], reverse=True)[:5],
            'upcoming_events': upcoming_events,
        }

        return Response(data)


class PartnerFestivalListView(APIView):
    """사업자용 축제 탐색"""
    permission_classes = [permissions.AllowAny]

    def get(self, request):
        # 모든 축제 (과거 포함)
        festivals = Event.objects.all().order_by('-start_date')

        # 검색 필터
        search = request.query_params.get('search', None)
        if search:
            festivals = festivals.filter(
                Q(name__icontains=search) |
                Q(location__icontains=search) |
                Q(description__icontains=search)
            )

        # 카테고리 필터
        category = request.query_params.get('category', None)
        if category:
            festivals = festivals.filter(category=category)

        # 지역 필터
        location = request.query_params.get('location', None)
        if location:
            festivals = festivals.filter(location__icontains=location)

        # 이미 지원한 축제 표시 (로그인한 경우만)
        applied_event_ids = []
        if request.user.is_authenticated and hasattr(request.user, 'partner_profile'):
            partner = request.user.partner_profile
            applied_event_ids = list(partner.applications.values_list('event_id', flat=True))

        from events.serializers import EventSerializer
        festivals_data = EventSerializer(festivals, many=True).data

        for festival in festivals_data:
            festival['already_applied'] = festival['id'] in applied_event_ids

        return Response({
            'count': festivals.count(),
            'results': festivals_data
        })


class ImageUploadViewSet(viewsets.ModelViewSet):
    """이미지 업로드 ViewSet"""
    queryset = ImageUpload.objects.all()
    serializer_class = ImageUploadSerializer
    permission_classes = [IsPartner]
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        """현재 사업자의 이미지만 반환"""
        return ImageUpload.objects.filter(
            partner__user=self.request.user
        ).select_related('partner')

    def perform_create(self, serializer):
        """이미지 업로드 생성 시 자동으로 partner 연결 및 리사이징"""
        partner = self.request.user.partner_profile
        image_file = self.request.FILES.get('image')

        # 이미지 리사이징 (선택적)
        resize = self.request.data.get('resize', 'true').lower() == 'true'
        if resize and image_file:
            resized_image = resize_image(
                image_file,
                max_width=1920,
                max_height=1080,
                quality=85
            )
            # 리사이징된 이미지로 교체
            self.request.FILES['image'] = resized_image

        serializer.save(partner=partner)

    @action(detail=False, methods=['post'], url_path='upload-certificate')
    def upload_certificate(self, request):
        """사업자등록증 업로드 전용 엔드포인트"""
        partner = request.user.partner_profile

        # 이미지 파일 확인
        if 'image' not in request.FILES:
            return Response(
                {'error': '이미지 파일이 필요합니다.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # 기존 사업자등록증이 있으면 삭제
        ImageUpload.objects.filter(
            partner=partner,
            image_type='certificate'
        ).delete()

        # 새 업로드 생성
        serializer = self.get_serializer(data={
            'image': request.FILES['image'],
            'image_type': 'certificate',
            'description': request.data.get('description', '사업자등록증')
        })

        if serializer.is_valid():
            serializer.save(partner=partner)
            return Response(serializer.data, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'], url_path='by-type/(?P<image_type>[^/.]+)')
    def by_type(self, request, image_type=None):
        """이미지 타입별 조회"""
        partner = request.user.partner_profile
        images = ImageUpload.objects.filter(
            partner=partner,
            image_type=image_type
        )
        serializer = self.get_serializer(images, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """업로드 통계"""
        partner = request.user.partner_profile
        images = ImageUpload.objects.filter(partner=partner)

        stats = {
            'total_count': images.count(),
            'total_size': sum(img.file_size for img in images),
            'by_type': {}
        }

        # 타입별 통계
        for image_type, _ in ImageUpload.IMAGE_TYPE_CHOICES:
            type_images = images.filter(image_type=image_type)
            stats['by_type'][image_type] = {
                'count': type_images.count(),
                'size': sum(img.file_size for img in type_images)
            }

        return Response(stats)


class NotificationViewSet(viewsets.ReadOnlyModelViewSet):
    """알림 ViewSet (읽기 전용)"""
    queryset = Notification.objects.all()
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        """현재 사용자의 알림만 반환"""
        return Notification.objects.filter(
            user=self.request.user
        ).select_related('application', 'application__event')

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        """알림 읽음 처리"""
        notification = self.get_object()
        notification.mark_as_read()
        serializer = self.get_serializer(notification)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        """모든 알림 읽음 처리"""
        notifications = self.get_queryset().filter(read=False)
        count = notifications.count()
        for notification in notifications:
            notification.mark_as_read()
        return Response({
            'message': f'{count}개의 알림을 읽음 처리했습니다.',
            'count': count
        })

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        """읽지 않은 알림 개수"""
        count = self.get_queryset().filter(read=False).count()
        return Response({'unread_count': count})

    @action(detail=False, methods=['get'])
    def unread(self, request):
        """읽지 않은 알림 목록"""
        notifications = self.get_queryset().filter(read=False)
        serializer = self.get_serializer(notifications, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'], url_path='by-type/(?P<notification_type>[^/.]+)')
    def by_type(self, request, notification_type=None):
        """타입별 알림 조회"""
        notifications = self.get_queryset().filter(notification_type=notification_type)
        serializer = self.get_serializer(notifications, many=True)
        return Response(serializer.data)


class ApplicationDraftViewSet(viewsets.ModelViewSet):
    """지원서 임시저장 ViewSet"""
    queryset = ApplicationDraft.objects.all()
    serializer_class = ApplicationDraftSerializer
    permission_classes = [IsPartner]

    def get_queryset(self):
        """현재 사업자의 임시저장만 반환"""
        return ApplicationDraft.objects.filter(
            partner__user=self.request.user
        ).select_related('partner', 'event')

    def create(self, request, *args, **kwargs):
        """임시저장 생성 (기존 것이 있으면 업데이트)"""
        partner = request.user.partner_profile
        event_id = request.data.get('event')

        # 기존 임시저장이 있는지 확인
        existing_draft = ApplicationDraft.objects.filter(
            partner=partner,
            event_id=event_id
        ).first()

        if existing_draft:
            # 기존 임시저장 업데이트
            serializer = self.get_serializer(existing_draft, data=request.data, partial=True)
            serializer.is_valid(raise_exception=True)
            self.perform_update(serializer)
            return Response(serializer.data)
        else:
            # 새로운 임시저장 생성
            return super().create(request, *args, **kwargs)

    def perform_create(self, serializer):
        """임시저장 생성 시 자동으로 partner 연결"""
        partner = self.request.user.partner_profile
        serializer.save(partner=partner)

    def perform_update(self, serializer):
        """업데이트 시에도 partner 유지"""
        partner = self.request.user.partner_profile
        serializer.save(partner=partner)

    @action(detail=False, methods=['get'], url_path='by-event/(?P<event_id>[^/.]+)')
    def by_event(self, request, event_id=None):
        """특정 축제의 임시저장 조회"""
        try:
            draft = self.get_queryset().get(event_id=event_id)
            serializer = self.get_serializer(draft)
            return Response(serializer.data)
        except ApplicationDraft.DoesNotExist:
            return Response({'detail': '임시저장이 없습니다.'}, status=status.HTTP_404_NOT_FOUND)


class FestivalBookmarkViewSet(viewsets.ModelViewSet):
    """축제 북마크 ViewSet"""
    queryset = FestivalBookmark.objects.all()
    serializer_class = FestivalBookmarkSerializer
    permission_classes = [IsPartner]

    def get_queryset(self):
        """현재 사업자의 북마크만 반환"""
        return FestivalBookmark.objects.filter(
            partner__user=self.request.user
        ).select_related('partner', 'event')

    def perform_create(self, serializer):
        """북마크 생성 시 자동으로 partner 연결"""
        partner = self.request.user.partner_profile
        serializer.save(partner=partner)

    @action(detail=False, methods=['post'], url_path='toggle/(?P<event_id>[^/.]+)')
    def toggle(self, request, event_id=None):
        """북마크 토글 (있으면 삭제, 없으면 생성)"""
        partner = request.user.partner_profile

        try:
            bookmark = FestivalBookmark.objects.get(partner=partner, event_id=event_id)
            bookmark.delete()
            return Response({'message': '북마크가 삭제되었습니다.', 'bookmarked': False})
        except FestivalBookmark.DoesNotExist:
            # 이벤트 존재 확인
            try:
                event = Event.objects.get(id=event_id)
            except Event.DoesNotExist:
                return Response({'error': '축제를 찾을 수 없습니다.'}, status=status.HTTP_404_NOT_FOUND)

            bookmark = FestivalBookmark.objects.create(partner=partner, event=event)
            serializer = self.get_serializer(bookmark)
            return Response({'message': '북마크가 추가되었습니다.', 'bookmarked': True, 'data': serializer.data}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['get'], url_path='check/(?P<event_id>[^/.]+)')
    def check(self, request, event_id=None):
        """특정 축제의 북마크 여부 확인"""
        partner = request.user.partner_profile
        bookmarked = FestivalBookmark.objects.filter(partner=partner, event_id=event_id).exists()
        return Response({'bookmarked': bookmarked})
